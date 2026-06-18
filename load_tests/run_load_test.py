"""
Civic Reporter — Load Test Runner
===================================
Runs Locust headlessly:  100 users | 1 minute
Then parses Locust's CSV output and produces a rich Excel report.
"""
import subprocess
import sys
import os
import csv
import json
import time
from datetime import datetime
from pathlib import Path

# ─── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
REPORTS_DIR  = SCRIPT_DIR / "reports"
CSV_PREFIX   = REPORTS_DIR / "load_test"
LOCUSTFILE   = SCRIPT_DIR / "locustfile.py"

# ─── Load-test parameters ─────────────────────────────────────────────────────
TARGET_URL   = os.environ.get("BASE_URL", "http://localhost:5000")
USERS        = 100      # concurrent virtual users
SPAWN_RATE   = 10       # users added per second until 100 are active
RUN_TIME     = "1m"     # duration: 1 minute

# ─── Install locust if missing ────────────────────────────────────────────────
def ensure_locust():
    try:
        import locust  # noqa
    except ImportError:
        print("📦  Locust not found — installing ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "locust", "-q"])
        print("✅  Locust installed.")

# ─── Run Locust (headless) ────────────────────────────────────────────────────
def run_locust():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    cmd = [
        sys.executable, "-m", "locust",
        "-f",            str(LOCUSTFILE),
        "--headless",
        "--host",        TARGET_URL,
        "--users",       str(USERS),
        "--spawn-rate",  str(SPAWN_RATE),
        "--run-time",    RUN_TIME,
        "--csv",         str(CSV_PREFIX),
        "--csv-full-history",
        "--only-summary",
    ]
    print(f"\n🚀  Starting load test against {TARGET_URL}")
    print(f"   Users: {USERS}  |  Spawn rate: {SPAWN_RATE}/s  |  Duration: {RUN_TIME}\n")
    result = subprocess.run(cmd, capture_output=False, text=True)
    if result.returncode not in (0, 1):   # Locust exits 1 even on success sometimes
        raise RuntimeError(f"Locust exited with code {result.returncode}")

# ─── Parse Locust CSV output ──────────────────────────────────────────────────
def parse_csv():
    stats_file = Path(str(CSV_PREFIX) + "_stats.csv")
    if not stats_file.exists():
        raise FileNotFoundError(f"Locust CSV not found: {stats_file}")

    rows = []
    with open(stats_file, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows

# ─── Build Excel report ───────────────────────────────────────────────────────
def build_excel(rows: list[dict]) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference

    # ── Style helpers ──────────────────────────────────────────────────────────
    BLUE      = "1A3A5C"
    BLUE_LT   = "2E6DA4"
    GREEN     = "27AE60"
    GREEN_LT  = "D5F5E3"
    RED       = "E74C3C"
    RED_LT    = "FADBD8"
    AMBER     = "F39C12"
    AMBER_LT  = "FDEBD0"
    WHITE     = "FFFFFF"
    GREY      = "F2F3F4"
    GREY_H    = "D5D8DC"
    TEAL      = "117A65"
    TEAL_LT   = "A2D9CE"

    def side():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(c):
        return PatternFill(start_color=c, end_color=c, fill_type="solid")

    def font(bold=False, size=10, color="222222", name="Calibri"):
        return Font(bold=bold, size=size, color=color, name=name)

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def left():
        return Alignment(horizontal="left", vertical="center")

    def hdr_cell(ws, row, col, value, bg=BLUE_LT, fg=WHITE, sz=10, bold=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font(bold=bold, size=sz, color=fg)
        c.fill      = fill(bg)
        c.alignment = center()
        c.border    = side()
        return c

    def data_cell(ws, row, col, value, bg=WHITE, fg="222222", bold=False, align="center"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font(bold=bold, size=9, color=fg)
        c.fill      = fill(bg)
        c.alignment = center() if align == "center" else left()
        c.border    = side()
        return c

    # ── Compute summary numbers ────────────────────────────────────────────────
    # Find "Aggregated" row
    agg = next((r for r in rows if r.get("Name", "").strip() == "Aggregated"), None)
    if agg is None and rows:
        agg = rows[-1]   # fallback: last row

    def safe_float(d, key, default=0.0):
        try: return float(d.get(key, default) or default)
        except: return default

    def safe_int(d, key, default=0):
        try: return int(float(d.get(key, default) or default))
        except: return default

    total_reqs      = safe_int(agg,   "Request Count")
    total_failures  = safe_int(agg,   "Failure Count")
    success_reqs    = total_reqs - total_failures
    avg_ms          = safe_float(agg, "Average Response Time")
    min_ms          = safe_float(agg, "Min Response Time")
    max_ms          = safe_float(agg, "Max Response Time")
    median_ms       = safe_float(agg, "Median Response Time")
    p90_ms          = safe_float(agg, "90%")
    p95_ms          = safe_float(agg, "95%")
    p99_ms          = safe_float(agg, "99%")
    rps             = safe_float(agg, "Requests/s")
    fail_rate       = round(total_failures / total_reqs * 100, 2) if total_reqs else 0
    success_rate    = round(100 - fail_rate, 2)

    # ── Health flag helpers ────────────────────────────────────────────────────
    def rt_colour(ms):
        """Green <500ms, Amber 500-1000ms, Red >1000ms"""
        if ms < 500:  return GREEN, GREEN_LT
        if ms < 1000: return AMBER, AMBER_LT
        return RED, RED_LT

    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = REPORTS_DIR / f"Civic_Reporter_LoadTest_{now_str}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 1 — Cover Page
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("🏙️ Cover Page")
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 30

    ws1.merge_cells("A1:C3")
    c = ws1["A1"]
    c.value     = "🏙️  CIVIC REPORTER — BASELINE LOAD TEST REPORT"
    c.font      = Font(bold=True, size=18, color=WHITE, name="Calibri")
    c.fill      = fill(BLUE)
    c.alignment = center()
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells("A4:C4")
    c = ws1["A4"]
    c.value     = f"100 Virtual Users  |  1 Minute  |  {datetime.now().strftime('%d %B %Y, %H:%M')}"
    c.font      = Font(bold=False, size=12, color=WHITE, name="Calibri")
    c.fill      = fill(BLUE_LT)
    c.alignment = center()
    ws1.row_dimensions[4].height = 22

    meta = [
        ("Target URL",        TARGET_URL),
        ("Test Type",         "Baseline / Load Test"),
        ("Concurrent Users",  USERS),
        ("Spawn Rate",        f"{SPAWN_RATE} users/sec"),
        ("Duration",          RUN_TIME),
        ("Test Executed",     datetime.now().strftime("%d %B %Y, %H:%M:%S")),
        ("Total Requests",    total_reqs),
        ("Successful",        f"{success_reqs}  ✅"),
        ("Failed",            f"{total_failures}  ❌"),
        ("Success Rate",      f"{success_rate}%"),
        ("Avg Response Time", f"{round(avg_ms, 1)} ms"),
        ("Requests/sec",      f"{round(rps, 2)} RPS"),
    ]
    for i, (label, value) in enumerate(meta, start=6):
        ws1.row_dimensions[i].height = 20
        lc = ws1.cell(row=i, column=2, value=label)
        lc.font      = font(bold=True, color="222222")
        lc.fill      = fill(GREY_H)
        lc.alignment = left()
        lc.border    = side()
        vc = ws1.cell(row=i, column=3, value=value)
        vc.font      = font(color="333333")
        vc.alignment = left()
        vc.border    = side()
        if label == "Successful":
            vc.fill = fill(GREEN_LT); vc.font = font(bold=True, color=GREEN)
        elif label == "Failed":
            vc.fill = fill(RED_LT);   vc.font = font(bold=True, color=RED)
        elif label == "Success Rate":
            vc.fill = fill(GREEN_LT if success_rate >= 95 else (AMBER_LT if success_rate >= 80 else RED_LT))
            vc.font = font(bold=True, color=GREEN if success_rate >= 95 else (AMBER if success_rate >= 80 else RED))

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 2 — Executive Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("📊 Executive Summary")
    for col, w in enumerate([22, 20, 20, 20, 20], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws2.merge_cells("A1:E1")
    c = ws2["A1"]
    c.value     = "EXECUTIVE SUMMARY — LOAD TEST RESULTS"
    c.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.fill      = fill(BLUE)
    c.alignment = center()
    ws2.row_dimensions[1].height = 28

    # KPI cards (row 3+)
    kpis = [
        ("Total Requests",   total_reqs,          "reqs",  BLUE_LT,  WHITE),
        ("Successful",       success_reqs,         "reqs",  GREEN,    WHITE),
        ("Failed",           total_failures,       "reqs",  RED,      WHITE),
        ("Success Rate",     f"{success_rate}%",  "",      GREEN if success_rate >= 95 else AMBER, WHITE),
        ("Req / Second",     f"{round(rps,2)}",   "RPS",   TEAL,     WHITE),
    ]
    ws2.merge_cells("A2:E2")
    ws2["A2"].value     = "📌  Key Performance Indicators"
    ws2["A2"].font      = font(bold=True, size=11, color=WHITE)
    ws2["A2"].fill      = fill(BLUE_LT)
    ws2["A2"].alignment = left()

    for col, (label, val, unit, bg, fg) in enumerate(kpis, 1):
        ws2.column_dimensions[get_column_letter(col)].width = 20
        hdr = ws2.cell(row=3, column=col, value=label)
        hdr.font = font(bold=True, size=9, color=WHITE); hdr.fill = fill(bg)
        hdr.alignment = center(); hdr.border = side()
        ws2.row_dimensions[3].height = 16

        val_cell = ws2.cell(row=4, column=col, value=val)
        val_cell.font      = Font(bold=True, size=18, color=bg, name="Calibri")
        val_cell.alignment = center()
        val_cell.border    = side()
        ws2.row_dimensions[4].height = 36

        u_cell = ws2.cell(row=5, column=col, value=unit)
        u_cell.font      = font(size=8, color="888888")
        u_cell.alignment = center()
        u_cell.border    = side()
        ws2.row_dimensions[5].height = 14

    # Response time table (row 7+)
    ws2.merge_cells("A7:E7")
    c = ws2["A7"]
    c.value     = "⏱️  Response Time Breakdown (milliseconds)"
    c.font      = font(bold=True, size=11, color=WHITE)
    c.fill      = fill(BLUE_LT)
    c.alignment = left()
    ws2.row_dimensions[7].height = 22

    rt_headers = ["Metric", "Min", "Average", "Median", "Max"]
    for col, h in enumerate(rt_headers, 1):
        hdr_cell(ws2, 8, col, h, BLUE_LT)
    ws2.row_dimensions[8].height = 18

    rt_rows = [
        ("Response Time (ms)", min_ms, avg_ms, median_ms, max_ms),
        ("Percentile p90",     "",     "",     "",        p90_ms),
        ("Percentile p95",     "",     "",     "",        p95_ms),
        ("Percentile p99",     "",     "",     "",        p99_ms),
    ]
    for row_i, (label, mn, av, med, mx) in enumerate(rt_rows, 9):
        data_cell(ws2, row_i, 1, label, GREY, bold=True, align="left")
        for col_i, val in enumerate([mn, av, med, mx], 2):
            if val == "": data_cell(ws2, row_i, col_i, "—"); continue
            fgc, bgc = rt_colour(val)
            data_cell(ws2, row_i, col_i, f"{round(val,1)} ms", bgc, fgc, bold=True)
        ws2.row_dimensions[row_i].height = 16

    # Threshold evaluation (row 15+)
    ws2.merge_cells("A15:E15")
    c = ws2["A15"]
    c.value     = "✅  Performance Thresholds"
    c.font      = font(bold=True, size=11, color=WHITE)
    c.fill      = fill(BLUE_LT)
    c.alignment = left()

    thresholds = [
        ("Avg Response Time < 500ms",   avg_ms < 500,       f"Actual: {round(avg_ms,1)}ms"),
        ("Max Response Time < 2000ms",  max_ms < 2000,      f"Actual: {round(max_ms,1)}ms"),
        ("p95 Response Time < 1000ms",  p95_ms < 1000,      f"Actual: {round(p95_ms,1)}ms"),
        ("Success Rate ≥ 95%",          success_rate >= 95, f"Actual: {success_rate}%"),
        ("RPS > 50",                    rps > 50,           f"Actual: {round(rps,2)} RPS"),
    ]
    th_headers = ["Threshold", "Target", "Actual / Result", "Status"]
    for col, h in enumerate(th_headers, 1):
        hdr_cell(ws2, 16, col, h, BLUE_LT)
    ws2.row_dimensions[16].height = 18

    for row_i, (desc, passed, actual) in enumerate(thresholds, 17):
        data_cell(ws2, row_i, 1, desc,                    GREY,     align="left")
        data_cell(ws2, row_i, 2, "Target",                GREY)
        data_cell(ws2, row_i, 3, actual,                  GREY)
        status_bg = GREEN_LT if passed else RED_LT
        status_fg = GREEN    if passed else RED
        data_cell(ws2, row_i, 4, "✅ PASS" if passed else "❌ FAIL",
                  status_bg, status_fg, bold=True)
        ws2.row_dimensions[row_i].height = 16

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 3 — Per-Endpoint Details
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("📝 Endpoint Details")
    col_widths = [28, 10, 10, 14, 14, 12, 12, 12, 12, 14]
    headers3   = ["Endpoint / Name", "Method", "Reqs", "Failures",
                  "Fail %", "Avg (ms)", "Min (ms)", "Max (ms)", "Med (ms)", "RPS"]
    for i, w in enumerate(col_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.merge_cells(f"A1:{get_column_letter(len(headers3))}1")
    c = ws3["A1"]
    c.value     = f"PER-ENDPOINT LOAD TEST RESULTS — {datetime.now().strftime('%d %b %Y %H:%M')}"
    c.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.fill      = fill(BLUE)
    c.alignment = center()
    ws3.row_dimensions[1].height = 26

    for col, h in enumerate(headers3, 1):
        hdr_cell(ws3, 2, col, h)
    ws3.row_dimensions[2].height = 18

    data_rows_ws3 = [r for r in rows if r.get("Name", "").strip() != "Aggregated"]
    for row_i, row in enumerate(data_rows_ws3, 3):
        reqs   = safe_int(row, "Request Count")
        fails  = safe_int(row, "Failure Count")
        fp     = round(fails / reqs * 100, 1) if reqs else 0
        avg_r  = safe_float(row, "Average Response Time")
        min_r  = safe_float(row, "Min Response Time")
        max_r  = safe_float(row, "Max Response Time")
        med_r  = safe_float(row, "Median Response Time")
        rps_r  = safe_float(row, "Requests/s")
        fail_c  = RED_LT  if fails > 0 else WHITE
        fail_fc = RED     if fails > 0 else "222222"

        data_cell(ws3, row_i, 1,  row.get("Name", ""), GREY, align="left")
        data_cell(ws3, row_i, 2,  row.get("Method", "GET"))
        data_cell(ws3, row_i, 3,  reqs)
        data_cell(ws3, row_i, 4,  fails,       fail_c, fail_fc, bold=(fails > 0))
        data_cell(ws3, row_i, 5,  f"{fp}%",    fail_c, fail_fc)
        _, bg_avg = rt_colour(avg_r)
        data_cell(ws3, row_i, 6,  f"{round(avg_r,1)}", bg_avg)
        data_cell(ws3, row_i, 7,  f"{round(min_r,1)}")
        data_cell(ws3, row_i, 8,  f"{round(max_r,1)}")
        data_cell(ws3, row_i, 9,  f"{round(med_r,1)}")
        data_cell(ws3, row_i, 10, f"{round(rps_r,2)}")
        ws3.row_dimensions[row_i].height = 16

    ws3.freeze_panes = "A3"

    # Footer aggregated row
    fr = len(data_rows_ws3) + 4
    ws3.merge_cells(f"A{fr}:{get_column_letter(len(headers3))}{fr}")
    fc = ws3[f"A{fr}"]
    fc.value     = (f"AGGREGATE  |  Total Reqs: {total_reqs}  |  "
                    f"Failed: {total_failures}  |  "
                    f"Avg: {round(avg_ms,1)}ms  |  RPS: {round(rps,2)}")
    fc.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
    fc.fill      = fill(BLUE)
    fc.alignment = center()
    ws3.row_dimensions[fr].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 4 — Charts
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("📈 Charts")
    ws4.merge_cells("A1:J1")
    c = ws4["A1"]
    c.value     = "LOAD TEST — VISUAL ANALYSIS"
    c.font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
    c.fill      = fill(BLUE)
    c.alignment = center()
    ws4.row_dimensions[1].height = 28

    # Build small helper table for the bar chart
    chart_headers = ["Endpoint", "Avg RT (ms)", "Max RT (ms)", "RPS"]
    for col, h in enumerate(chart_headers, 1):
        hdr_cell(ws4, 3, col, h, BLUE_LT)
    chart_data_rows = [r for r in rows if r.get("Name","").strip() != "Aggregated"]
    for row_i, row in enumerate(chart_data_rows, 4):
        ws4.cell(row=row_i, column=1, value=row.get("Name","")).alignment = left()
        ws4.cell(row=row_i, column=2, value=safe_float(row, "Average Response Time"))
        ws4.cell(row=row_i, column=3, value=safe_float(row, "Max Response Time"))
        ws4.cell(row=row_i, column=4, value=safe_float(row, "Requests/s"))
        ws4.row_dimensions[row_i].height = 14

    num_ep = len(chart_data_rows)
    if num_ep > 0:
        # Bar chart — Avg vs Max response times
        bar = BarChart()
        bar.type        = "col"
        bar.title       = "Response Time per Endpoint (ms)"
        bar.y_axis.title = "Milliseconds"
        bar.x_axis.title = "Endpoint"
        bar.width, bar.height = 20, 12
        data_ref = Reference(ws4, min_col=2, max_col=3, min_row=3, max_row=3 + num_ep)
        cats_ref = Reference(ws4, min_col=1, min_row=4, max_row=3 + num_ep)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        ws4.add_chart(bar, "A5")

        # RPS bar chart
        rps_bar = BarChart()
        rps_bar.type        = "col"
        rps_bar.title       = "Requests per Second (RPS) per Endpoint"
        rps_bar.y_axis.title = "RPS"
        rps_bar.width, rps_bar.height = 20, 12
        rps_data = Reference(ws4, min_col=4, max_col=4, min_row=3, max_row=3 + num_ep)
        rps_bar.add_data(rps_data, titles_from_data=True)
        rps_bar.set_categories(cats_ref)
        ws4.add_chart(rps_bar, "L5")

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 5 — Pass/Fail Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("✅ Pass-Fail Summary")
    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 18
    ws5.column_dimensions["C"].width = 20

    ws5.merge_cells("A1:C1")
    c = ws5["A1"]
    c.value     = "PASS / FAIL SUMMARY"
    c.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.fill      = fill(BLUE)
    c.alignment = center()
    ws5.row_dimensions[1].height = 26

    for col, h in enumerate(["Threshold Check", "Result", "Notes"], 1):
        hdr_cell(ws5, 2, col, h)

    all_checks = thresholds   # reuse from earlier
    for row_i, (desc, passed, actual) in enumerate(all_checks, 3):
        data_cell(ws5, row_i, 1, desc, GREY, align="left")
        status_bg = GREEN_LT if passed else RED_LT
        status_fg = GREEN    if passed else RED
        data_cell(ws5, row_i, 2, "✅ PASS" if passed else "❌ FAIL",
                  status_bg, status_fg, bold=True)
        data_cell(ws5, row_i, 3, actual, GREY)
        ws5.row_dimensions[row_i].height = 18

    # Overall verdict
    all_passed = all(p for _, p, _ in all_checks)
    vr = len(all_checks) + 4
    ws5.merge_cells(f"A{vr}:C{vr}")
    vc = ws5[f"A{vr}"]
    vc.value     = ("🎉  ALL CHECKS PASSED — System performs within acceptable thresholds!"
                    if all_passed else
                    "⚠️  SOME CHECKS FAILED — Review response times and error rates!")
    vc.font      = Font(bold=True, size=12, color=WHITE, name="Calibri")
    vc.fill      = fill(GREEN if all_passed else RED)
    vc.alignment = center()
    ws5.row_dimensions[vr].height = 28

    wb.save(filename)
    return str(filename)


# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("="*60)
    print(" CIVIC REPORTER — BASELINE LOAD TEST")
    print(f" Target: {TARGET_URL}")
    print(f" Users:  {USERS}  |  Duration: {RUN_TIME}")
    print("="*60)

    ensure_locust()
    print("\n⏱️  Running load test — please wait ~1 minute ...\n")
    run_locust()

    print("\n📊  Parsing results and generating Excel report ...")
    rows = parse_csv()
    report_path = build_excel(rows)

    print("\n" + "="*60)
    print(f"✅  DONE!  Report saved to:\n   {report_path}")
    print("="*60)
