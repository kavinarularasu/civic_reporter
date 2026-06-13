"""
Excel Report Generator — Rich multi-sheet analysis report with charts
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
import config


# ────────────────────────────────────────────────────────────
# Colour palette (matching Civic Reporter brand)
# ────────────────────────────────────────────────────────────
BRAND_BLUE    = "1A5276"
BRAND_BLUE_LT = "2E86C1"
GREEN_PASS    = "27AE60"
GREEN_LT      = "ABEBC6"
RED_FAIL      = "E74C3C"
RED_LT        = "F5B7B1"
AMBER_SKIP    = "F39C12"
AMBER_LT      = "FAD7A0"
WHITE         = "FFFFFF"
GREY_BG       = "F2F3F4"
GREY_HEADER   = "D5D8DC"

# Border helper
def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, size=10, color=WHITE, name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


class ExcelReportGenerator:
    """
    Generates a rich Excel test report with:
      Sheet 1 — Cover Page
      Sheet 2 — Executive Summary (with bar + pie charts)
      Sheet 3 — Detailed Test Results
      Sheet 4 — Module-wise Analysis
      Sheet 5 — Failure Analysis
      Sheet 6 — Screenshots Index
    """

    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or config.REPORT_DIR
        os.makedirs(self.output_dir, exist_ok=True)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)   # Remove default blank sheet

    # ────────────────────────────────────────────
    # Public API
    # ────────────────────────────────────────────
    def generate_report(
        self,
        test_results: list[dict],
        filename: str = None,
        device_info: dict = None
    ) -> str:
        """Build all sheets and save the workbook. Returns saved path."""
        filename = filename or config.EXCEL_REPORT_NAME
        stats    = self._calc_stats(test_results)
        info     = device_info or {}

        self._sheet_cover(stats, info)
        self._sheet_summary(test_results, stats)
        self._sheet_results(test_results, stats)
        self._sheet_module_analysis(test_results)
        self._sheet_failure_analysis(test_results)
        self._sheet_screenshots(test_results)

        path = os.path.join(self.output_dir, filename)
        self.wb.save(path)
        return path

    # ────────────────────────────────────────────
    # Statistics helper
    # ────────────────────────────────────────────
    @staticmethod
    def _calc_stats(results: list[dict]) -> dict:
        total   = len(results)
        passed  = sum(1 for r in results if r.get("status") == "PASSED")
        failed  = sum(1 for r in results if r.get("status") == "FAILED")
        skipped = sum(1 for r in results if r.get("status") == "SKIPPED")
        rate    = round(passed / total * 100, 2) if total else 0
        return dict(total=total, passed=passed, failed=failed,
                    skipped=skipped, rate=rate)

    @staticmethod
    def _module_of(test_name: str) -> str:
        """Extract module prefix from test name, e.g. TC01 → Auth."""
        n = test_name.upper()
        if n.startswith("TC0") or n.startswith("TC1"):
            tc = int("".join(filter(str.isdigit, test_name.split("_")[0])) or "0")
            if  1 <= tc <= 10:  return "01-Splash & Login"
            if 11 <= tc <= 22:  return "02-Home Dashboard"
            if 23 <= tc <= 36:  return "03-Report Issue"
            if 37 <= tc <= 49:  return "04-My Reports"
            if 50 <= tc <= 59:  return "05-Map Screen"
            if 60 <= tc <= 70:  return "06-Profile & Notif"
            if 71 <= tc <= 86:  return "07-Officer Portal"
        if n.startswith("E2E"): return "08-E2E Journey"
        return "General"

    # ────────────────────────────────────────────
    # Sheet 1 — Cover Page
    # ────────────────────────────────────────────
    def _sheet_cover(self, stats, info):
        ws = self.wb.create_sheet("📋 Cover Page")
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 35

        # Title banner
        ws.merge_cells("A1:C3")
        cell = ws["A1"]
        cell.value     = "🏙️  CIVIC REPORTER — MOBILE TEST REPORT"
        cell.font      = Font(bold=True, size=18, color=WHITE, name="Calibri")
        cell.fill      = _fill(BRAND_BLUE)
        cell.alignment = _center()

        # Sub-title
        ws.merge_cells("A4:C4")
        cell = ws["A4"]
        cell.value     = "Appium End-to-End Automation | Android"
        cell.font      = Font(bold=False, size=12, color=WHITE, name="Calibri")
        cell.fill      = _fill(BRAND_BLUE_LT)
        cell.alignment = _center()

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[4].height = 22

        rows = [
            ("Generated",       datetime.now().strftime("%d %B %Y, %H:%M")),
            ("App Package",     config.APP_PACKAGE),
            ("Device",          info.get("device",   config.ANDROID_DEVICE_NAME)),
            ("Android Version", info.get("version",  config.ANDROID_PLATFORM_VERSION)),
            ("Appium Server",   config.APPIUM_URL),
            ("Total Tests",     stats["total"]),
            ("Passed",          f"{stats['passed']}  ✅"),
            ("Failed",          f"{stats['failed']}  ❌"),
            ("Skipped",         f"{stats['skipped']}  ⚠️"),
            ("Pass Rate",       f"{stats['rate']}%"),
        ]

        for i, (label, value) in enumerate(rows, start=6):
            ws.row_dimensions[i].height = 20
            lc = ws.cell(row=i, column=2, value=label)
            lc.font      = _font(bold=True, color="222222")
            lc.fill      = _fill(GREY_HEADER)
            lc.alignment = _left()
            lc.border    = _thin_border()

            vc = ws.cell(row=i, column=3, value=value)
            vc.font      = _font(bold=False, color="333333")
            vc.alignment = _left()
            vc.border    = _thin_border()

            # Colour pass/fail rows
            if label == "Passed":
                vc.fill = _fill(GREEN_LT); vc.font = _font(bold=True, color=GREEN_PASS)
            elif label == "Failed":
                vc.fill = _fill(RED_LT);   vc.font = _font(bold=True, color=RED_FAIL)
            elif label == "Pass Rate":
                rate_val = stats["rate"]
                vc.fill  = _fill(GREEN_LT if rate_val >= 80 else RED_LT)
                vc.font  = _font(bold=True, color=GREEN_PASS if rate_val >= 80 else RED_FAIL)

    # ────────────────────────────────────────────
    # Sheet 2 — Executive Summary + Charts
    # ────────────────────────────────────────────
    def _sheet_summary(self, results, stats):
        ws = self.wb.create_sheet("📊 Summary")
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        # Header
        ws.merge_cells("A1:D1")
        c = ws["A1"]
        c.value     = "EXECUTIVE SUMMARY"
        c.font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
        c.fill      = _fill(BRAND_BLUE)
        c.alignment = _center()
        ws.row_dimensions[1].height = 28

        # Stats table
        headers = ["Metric", "Count", "Percentage", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = _font(bold=True, color=WHITE)
            cell.fill      = _fill(BRAND_BLUE_LT)
            cell.alignment = _center()
            cell.border    = _thin_border()

        data_rows = [
            ("Total Tests",  stats["total"],   "100.00%",                           "🧪"),
            ("✅ Passed",    stats["passed"],   f"{stats['rate']}%",                 "PASS"),
            ("❌ Failed",    stats["failed"],   f"{round(stats['failed']/stats['total']*100,2) if stats['total'] else 0}%", "FAIL"),
            ("⚠️ Skipped",  stats["skipped"],  f"{round(stats['skipped']/stats['total']*100,2) if stats['total'] else 0}%","SKIP"),
        ]

        fill_map = {
            "PASS": (GREEN_LT, GREEN_PASS),
            "FAIL": (RED_LT,   RED_FAIL),
            "SKIP": (AMBER_LT, AMBER_SKIP),
            "🧪":   (GREY_BG,  "222222"),
        }

        for r, (metric, count, pct, tag) in enumerate(data_rows, 3):
            bg, fg = fill_map.get(tag, (WHITE, "333333"))
            values = [metric, count, pct, tag if tag not in fill_map else ""]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.alignment = _center()
                cell.border    = _thin_border()
                cell.fill      = _fill(bg)
                cell.font      = Font(bold=(col == 1), size=10, color=fg, name="Calibri")
            ws.row_dimensions[r].height = 18

        # ── Bar Chart ──────────────────────────
        bar = BarChart()
        bar.type       = "col"
        bar.title      = "Test Results Overview"
        bar.y_axis.title = "Count"
        bar.x_axis.title = "Status"
        bar.width, bar.height = 14, 10

        data_ref = Reference(ws, min_col=2, max_col=2, min_row=3, max_row=5)
        cats_ref = Reference(ws, min_col=1,             min_row=3, max_row=5)
        bar.add_data(data_ref)
        bar.set_categories(cats_ref)
        bar.series[0].graphicalProperties.solidFill = GREEN_PASS
        ws.add_chart(bar, "A8")

        # ── Pie Chart ──────────────────────────
        pie = PieChart()
        pie.title  = "Pass / Fail Distribution"
        pie.width, pie.height = 14, 10

        pie.add_data(Reference(ws, min_col=2, max_col=2, min_row=3, max_row=5))
        pie.set_categories(Reference(ws, min_col=1, min_row=3, max_row=5))
        colors = [GREEN_PASS, RED_FAIL, AMBER_SKIP]
        for i, color in enumerate(colors):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            pie.series[0].dPt.append(pt)
        ws.add_chart(pie, "E8")

    # ────────────────────────────────────────────
    # Sheet 3 — Detailed Test Results
    # ────────────────────────────────────────────
    def _sheet_results(self, results, stats):
        ws = self.wb.create_sheet("📝 Test Results")
        col_widths = [6, 45, 12, 22, 20, 55, 35]
        headers    = ["#", "Test Name", "Status", "Module", "Timestamp", "Details", "Screenshot"]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Header row
        ws.row_dimensions[1].height = 24
        ws.merge_cells("A1:G1")
        cell = ws["A1"]
        cell.value     = f"DETAILED TEST RESULTS — {datetime.now().strftime('%d %b %Y %H:%M')}"
        cell.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
        cell.fill      = _fill(BRAND_BLUE)
        cell.alignment = _center()

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = _font(bold=True, color=WHITE)
            cell.fill      = _fill(BRAND_BLUE_LT)
            cell.alignment = _center()
            cell.border    = _thin_border()
        ws.row_dimensions[2].height = 20

        for row_i, res in enumerate(results, 3):
            status = res.get("status", "UNKNOWN")
            module = self._module_of(res.get("test_name", ""))

            if   status == "PASSED":  bg, fg = GREEN_LT, GREEN_PASS
            elif status == "FAILED":  bg, fg = RED_LT,   RED_FAIL
            elif status == "SKIPPED": bg, fg = AMBER_LT, AMBER_SKIP
            else:                     bg, fg = GREY_BG,  "333333"

            row_vals = [
                row_i - 2,
                res.get("test_name",  ""),
                status,
                module,
                res.get("timestamp",  ""),
                res.get("details",    ""),
                os.path.basename(res.get("screenshot", "")) or "—",
            ]
            for col, val in enumerate(row_vals, 1):
                cell            = ws.cell(row=row_i, column=col, value=val)
                cell.border     = _thin_border()
                cell.alignment  = _center(wrap=(col in (2, 6)))
                cell.fill       = _fill(WHITE)
                cell.font       = Font(size=9, name="Calibri", color="222222")
                # Status cell gets colour
                if col == 3:
                    cell.fill = _fill(bg)
                    cell.font = Font(bold=True, size=9, color=fg, name="Calibri")
            ws.row_dimensions[row_i].height = 16

        # Freeze header rows
        ws.freeze_panes = "A3"

        # Summary footer
        footer_row = len(results) + 4
        ws.merge_cells(f"A{footer_row}:G{footer_row}")
        fc = ws[f"A{footer_row}"]
        fc.value     = (f"Total: {stats['total']}  |  Passed: {stats['passed']}  |  "
                        f"Failed: {stats['failed']}  |  Pass Rate: {stats['rate']}%")
        fc.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
        fc.fill      = _fill(BRAND_BLUE)
        fc.alignment = _center()

    # ────────────────────────────────────────────
    # Sheet 4 — Module-wise Analysis
    # ────────────────────────────────────────────
    def _sheet_module_analysis(self, results):
        ws = self.wb.create_sheet("📦 Module Analysis")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14

        ws.merge_cells("A1:E1")
        cell = ws["A1"]
        cell.value     = "MODULE-WISE TEST ANALYSIS"
        cell.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
        cell.fill      = _fill(BRAND_BLUE)
        cell.alignment = _center()
        ws.row_dimensions[1].height = 24

        headers = ["Module", "Total", "Passed", "Failed", "Pass Rate"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = _font(bold=True, color=WHITE)
            cell.fill      = _fill(BRAND_BLUE_LT)
            cell.alignment = _center()
            cell.border    = _thin_border()
        ws.row_dimensions[2].height = 20

        # Group by module
        modules: dict[str, dict] = {}
        for res in results:
            mod = self._module_of(res.get("test_name", ""))
            if mod not in modules:
                modules[mod] = {"total": 0, "passed": 0, "failed": 0}
            modules[mod]["total"] += 1
            if res.get("status") == "PASSED":
                modules[mod]["passed"] += 1
            elif res.get("status") == "FAILED":
                modules[mod]["failed"] += 1

        for row_i, (mod, d) in enumerate(sorted(modules.items()), 3):
            rate = round(d["passed"] / d["total"] * 100, 1) if d["total"] else 0
            bg   = GREEN_LT if rate == 100 else (RED_LT if rate < 50 else AMBER_LT)
            vals = [mod, d["total"], d["passed"], d["failed"], f"{rate}%"]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.alignment = _center()
                cell.border    = _thin_border()
                cell.fill      = _fill(GREY_BG if col == 1 else (bg if col == 5 else WHITE))
                cell.font      = Font(bold=(col == 1 or col == 5), size=9,
                                      name="Calibri", color="222222")
            ws.row_dimensions[row_i].height = 16

        # Bar chart for module pass rates
        chart_row = len(modules) + 5
        bar = BarChart()
        bar.type         = "col"
        bar.title        = "Module-wise Pass Rate (%)"
        bar.y_axis.title = "Pass Rate %"
        bar.width, bar.height = 20, 12

        data_ref = Reference(ws, min_col=5, max_col=5,
                             min_row=2, max_row=2 + len(modules))
        cats_ref = Reference(ws, min_col=1,
                             min_row=3, max_row=2 + len(modules))
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        ws.add_chart(bar, f"A{chart_row}")

    # ────────────────────────────────────────────
    # Sheet 5 — Failure Analysis
    # ────────────────────────────────────────────
    def _sheet_failure_analysis(self, results):
        ws = self.wb.create_sheet("❌ Failures")
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 60

        ws.merge_cells("A1:E1")
        cell = ws["A1"]
        cell.value     = "FAILURE ANALYSIS"
        cell.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
        cell.fill      = _fill(RED_FAIL)
        cell.alignment = _center()
        ws.row_dimensions[1].height = 24

        failures = [r for r in results if r.get("status") == "FAILED"]
        if not failures:
            ws.merge_cells("A2:E2")
            fc = ws["A2"]
            fc.value     = "🎉  All tests PASSED — No failures recorded!"
            fc.font      = Font(bold=True, size=12, color=GREEN_PASS, name="Calibri")
            fc.fill      = _fill(GREEN_LT)
            fc.alignment = _center()
            return

        headers = ["#", "Test Name", "Module", "Timestamp", "Error Detail"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = _font(bold=True, color=WHITE)
            cell.fill      = _fill(RED_FAIL)
            cell.alignment = _center()
            cell.border    = _thin_border()
        ws.row_dimensions[2].height = 20

        for row_i, res in enumerate(failures, 3):
            vals = [
                row_i - 2,
                res.get("test_name",  ""),
                self._module_of(res.get("test_name", "")),
                res.get("timestamp",  ""),
                res.get("details",    ""),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.alignment = _center(wrap=(col == 5))
                cell.border    = _thin_border()
                cell.fill      = _fill(RED_LT if row_i % 2 == 0 else WHITE)
                cell.font      = Font(size=9, name="Calibri", color="222222")
            ws.row_dimensions[row_i].height = 28

    # ────────────────────────────────────────────
    # Sheet 6 — Screenshots Index
    # ────────────────────────────────────────────
    def _sheet_screenshots(self, results):
        ws = self.wb.create_sheet("📸 Screenshots")
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 55

        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value     = "SCREENSHOTS INDEX"
        cell.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
        cell.fill      = _fill(BRAND_BLUE)
        cell.alignment = _center()
        ws.row_dimensions[1].height = 24

        headers = ["#", "Test Name", "Status", "Screenshot Path"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = _font(bold=True, color=WHITE)
            cell.fill      = _fill(BRAND_BLUE_LT)
            cell.alignment = _center()
            cell.border    = _thin_border()
        ws.row_dimensions[2].height = 20

        for row_i, res in enumerate(results, 3):
            shot = res.get("screenshot", "")
            status = res.get("status", "")
            bg = GREEN_LT if status == "PASSED" else (RED_LT if status == "FAILED" else WHITE)
            vals = [row_i - 2, res.get("test_name", ""), status, shot or "—"]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_i, column=col, value=val)
                cell.alignment = _center()
                cell.border    = _thin_border()
                cell.fill      = _fill(bg if col == 3 else WHITE)
                cell.font      = Font(size=9, name="Calibri", color="222222")
            ws.row_dimensions[row_i].height = 15
